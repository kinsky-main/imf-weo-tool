from __future__ import annotations

from pathlib import Path
from uuid import uuid4

import pandas as pd
from openpyxl import load_workbook

from weo_tools.app import enrich_for_legacy_columns, pivot_for_excel, run_excel_export, save_dataframe
from weo_tools.configuration import RuntimeSettings
from weo_tools.imf import AvailabilityAggregate, AvailabilityResult, Catalog, ReleaseInfo, _build_dataframe


class ExcelClient:
    def fetch_catalog(self) -> Catalog:
        return Catalog(
            release=ReleaseInfo(version="9.0.0", updated_at="2026-03-12", name="WEO"),
            countries={"GBR": "United Kingdom"},
            country_groups={},
            locations={"GBR": "United Kingdom"},
            indicators={"NGDPD": "Gross domestic product, current prices"},
            units={"USD": "U.S. dollars"},
            scales={"9": "Billions"},
        )

    def fetch_indicator_availability(
        self,
        values: list[str],
        frequency: str,
        *,
        strict: bool = True,
    ) -> AvailabilityAggregate:
        del frequency, strict
        return AvailabilityAggregate(
            results=[AvailabilityResult(requested_code=value, available_codes=["NGDPD"], series_count=1) for value in values],
            available_codes=["NGDPD"],
            common_codes=["NGDPD"],
            counts_by_code={"NGDPD": len(values)},
        )

    def fetch_dataframe(self, **_kwargs) -> pd.DataFrame:
        return pd.DataFrame(
            [
                {
                    "Country": "United Kingdom",
                    "Subject Descriptor": "Gross domestic product, current prices",
                    "Units": "U.S. dollars",
                    "Scale": "Billions",
                    "time_period": 2024,
                    "obs_value": 3644.636,
                },
                {
                    "Country": "United Kingdom",
                    "Subject Descriptor": "Gross domestic product, current prices",
                    "Units": "U.S. dollars",
                    "Scale": "Billions",
                    "time_period": 2025,
                    "obs_value": pd.NA,
                },
            ]
        )


def test_build_dataframe_normalizes_sdmx_response() -> None:
    csv_frame = pd.DataFrame(
        [
            {
                "COUNTRY": "GBR",
                "INDICATOR": "NGDPD",
                "FREQUENCY": "A",
                "TIME_PERIOD": "2024",
                "OBS_VALUE": "3644636000000",
                "SCALE": "9",
                "COUNTRY_UPDATE_DATE": "9/13/2025",
            },
            {
                "COUNTRY": "GBR",
                "INDICATOR": "NGDPD",
                "FREQUENCY": "A",
                "TIME_PERIOD": "2025",
                "OBS_VALUE": "3958780000000",
                "SCALE": "9",
                "COUNTRY_UPDATE_DATE": "9/13/2025",
            },
        ]
    )

    dataframe = _build_dataframe(
        csv_frame=csv_frame,
        dataset_version="9.0.0",
        country_labels={"GBR": "United Kingdom"},
        subject_labels={"NGDPD": "Gross domestic product, current prices"},
        unit_labels={"USD": "U.S. dollars"},
        scale_labels={"9": "Billions"},
        indicator_unit_labels={"NGDPD": "U.S. dollars"},
        indicator_unit_codes={"NGDPD": "USD"},
        indicator_scale_labels={"NGDPD": "Billions"},
    )
    dataframe = enrich_for_legacy_columns(dataframe)

    assert list(dataframe["Country"].unique()) == ["United Kingdom"]
    assert list(dataframe["Subject Descriptor"].unique()) == ["Gross domestic product, current prices"]
    assert list(dataframe["Units"].unique()) == ["U.S. dollars"]
    assert list(dataframe["Scale"].unique()) == ["Billions"]
    assert dataframe["obs_value"].tolist() == [3644636000000.0, 3958780000000.0]


def test_pivot_for_excel_creates_wide_year_columns() -> None:
    dataframe = pd.DataFrame(
        [
            {
                "Country": "United Kingdom",
                "Subject Descriptor": "Gross domestic product, current prices",
                "Units": "U.S. dollars",
                "Scale": "Billions",
                "time_period": 2024,
                "obs_value": 3644.636,
            },
            {
                "Country": "United Kingdom",
                "Subject Descriptor": "Gross domestic product, current prices",
                "Units": "U.S. dollars",
                "Scale": "Billions",
                "time_period": 2025,
                "obs_value": 3958.78,
            },
        ]
    )

    wide = pivot_for_excel(dataframe)

    assert list(wide.columns) == ["Country", "Subject Descriptor", "Units", "Scale", "2024", "2025"]
    assert wide.iloc[0]["2024"] == 3644.636


def _output_path(name: str) -> Path:
    directory = Path(".tmp") / "test_outputs"
    directory.mkdir(parents=True, exist_ok=True)
    return directory / f"{uuid4().hex}_{name}"


def test_save_dataframe_writes_numeric_xlsx_cells() -> None:
    dataframe = pd.DataFrame(
        [
            {"time_period": "2024", "obs_value": "3644.636", "country": "United Kingdom"},
            {"time_period": pd.NA, "obs_value": pd.NA, "country": "Austria"},
        ]
    )

    path = save_dataframe(dataframe, _output_path("frame.xlsx"))

    workbook = load_workbook(path)
    sheet = workbook.active

    assert sheet["A2"].data_type == "n"
    assert sheet["B2"].data_type == "n"
    assert sheet["A2"].value == 2024
    assert sheet["B2"].value == 3644.636


def test_run_excel_export_writes_numeric_year_cells(monkeypatch) -> None:
    settings = RuntimeSettings(
        countries=["GBR"],
        subject_descriptors=["NGDPD"],
        output_path=str(_output_path("weo_export.xlsx")),
    )
    client = ExcelClient()

    monkeypatch.setattr("weo_tools.app.run_with_status", lambda message, func, /, *args, **kwargs: func(*args, **kwargs))

    path = run_excel_export(settings, client)

    workbook = load_workbook(path)
    sheet = workbook["WEO Data"]

    assert sheet["E2"].data_type == "n"
    assert sheet["F2"].value is None
    assert sheet["E2"].value == 3644.636

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import re
import tomllib

from openpyxl import load_workbook


def normalize_label(value: str) -> str:
    compact = re.sub(r"[^0-9a-z]+", " ", value.lower()).strip()
    return re.sub(r"\s+", " ", compact)


@dataclass(slots=True)
class AliasConfig:
    units: dict[str, str] = field(default_factory=dict)
    scales: dict[str, str] = field(default_factory=dict)
    unit_display: dict[str, str] = field(default_factory=dict)
    scale_display: dict[str, str] = field(default_factory=dict)
    countries: dict[str, str] = field(default_factory=dict)
    subjects: dict[str, str] = field(default_factory=dict)


@dataclass(slots=True)
class LegacyCatalog:
    country_aliases: dict[str, set[str]] = field(default_factory=dict)
    subject_aliases: dict[str, set[str]] = field(default_factory=dict)
    preferred_country_labels: dict[str, str] = field(default_factory=dict)
    preferred_subject_labels: dict[str, str] = field(default_factory=dict)
    indicator_units: dict[str, set[str]] = field(default_factory=dict)
    indicator_scales: dict[str, set[str]] = field(default_factory=dict)
    preferred_unit_labels: dict[str, str] = field(default_factory=dict)
    preferred_scale_labels: dict[str, str] = field(default_factory=dict)


def load_alias_config(path: str | Path) -> AliasConfig:
    alias_path = Path(path)
    if not alias_path.exists():
        return AliasConfig()

    with alias_path.open("rb") as handle:
        raw = tomllib.load(handle)

    def normalize_mapping(section: str) -> dict[str, str]:
        entries = raw.get(section, {})
        return {normalize_label(key): str(value) for key, value in entries.items()}

    return AliasConfig(
        units=normalize_mapping("units"),
        scales=normalize_mapping("scales"),
        countries=normalize_mapping("countries"),
        subjects=normalize_mapping("subjects"),
        unit_display={str(key): str(value) for key, value in raw.get("unit_display", {}).items()},
        scale_display={str(key): str(value) for key, value in raw.get("scale_display", {}).items()},
    )


def load_legacy_catalog(path: str | Path) -> LegacyCatalog:
    workbook_path = Path(path)
    if not workbook_path.exists():
        return LegacyCatalog()

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    header_index = {str(value): idx for idx, value in enumerate(headers) if value is not None}

    country_idx = header_index["Country"]
    iso_idx = header_index["ISO"]
    subject_idx = header_index["Subject Descriptor"]
    subject_code_idx = header_index["WEO Subject Code"]
    units_idx = header_index["Units"]
    scale_idx = header_index["Scale"]

    catalog = LegacyCatalog()
    for row in rows:
        country_code = _clean(row[iso_idx])
        country_label = _clean(row[country_idx])
        subject_code = _clean(row[subject_code_idx])
        subject_label = _clean(row[subject_idx])
        unit_label = _clean(row[units_idx])
        scale_label = _clean(row[scale_idx])
        if country_code and country_label:
            catalog.country_aliases.setdefault(country_code, set()).add(country_label)
            catalog.preferred_country_labels.setdefault(country_code, country_label)
        if subject_code and subject_label:
            catalog.subject_aliases.setdefault(subject_code, set()).add(subject_label)
            catalog.preferred_subject_labels.setdefault(subject_code, subject_label)
        if subject_code and unit_label:
            catalog.indicator_units.setdefault(subject_code, set()).add(unit_label)
            catalog.preferred_unit_labels.setdefault(subject_code, unit_label)
        if subject_code and scale_label:
            catalog.indicator_scales.setdefault(subject_code, set()).add(scale_label)
            catalog.preferred_scale_labels.setdefault(subject_code, scale_label)

    workbook.close()
    return catalog


def _clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
